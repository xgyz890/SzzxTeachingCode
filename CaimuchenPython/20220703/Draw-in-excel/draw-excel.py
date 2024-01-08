from PIL import Image
import openpyxl
from openpyxl.styles import fills
from openpyxl.utils import get_column_letter
import os

MAX_WIDTH = 300
MAX_HEIGHT = 300

"""
问题：最大宽度和最大高度，设置为较大数值，比如600，生成的excel打开需要修复，修复后颜色消失。
解决方法：最大宽度和最大高度设置为300.
原因：
excel 的原形是由多个 xml 文件，填充的颜色都存储在一个style.xml文件里面，当这个文件过大就会导致打开的时候报错。
所以为了解决这个问题，有两个解决方案，第一是缩小图片，第二是减少图片颜色。缩小图片的时候自带减少图片颜色的功能，减少图片颜色的方法可以采用灰度化、二值化等方法。
总体上来讲，就是需要控制颜色数量*单元格数<阈值（3300w左右)
"""

def resize(img):
    w, h = img.size
    if w > MAX_WIDTH:
        h = MAX_WIDTH / w * h
        w = MAX_WIDTH

    if h > MAX_HEIGHT:
        w = MAX_HEIGHT / h * w
        h = MAX_HEIGHT
    return img.resize((int(w), int(h)), Image.ANTIALIAS)


def int_to_16(num):
    num1 = hex(num).replace('0x', '')
    num2 = num1 if len(num1) > 1 else '0' + num1
    return num2


def draw_jpg(img_path,excel_path):

    img_pic = resize(Image.open(img_path))
    if os.path.exists(excel_path):
        os.remove(excel_path)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    width, height = img_pic.size
    cell_width, cell_height = 1.0, 6.0 

    for w in range(0, width):
        for h in range(0, height):
            if img_pic.mode == 'RGB':
                r, g, b = img_pic.getpixel((w, h))
            elif img_pic.mode == 'RGBA':
                r, g, b, a = img_pic.getpixel((w, h))
            
            hex_rgb = int_to_16(r) + int_to_16(g) + int_to_16(b)
            cell = worksheet.cell(column=w+1, row=h+1)
            cell.fill = fills.PatternFill(fill_type="solid", fgColor=hex_rgb)

        print('write in:', w+1, '  |  all:', width)
    
    for i in range(1, worksheet.max_row + 1): 
        worksheet.row_dimensions[i].height = cell_height 
        
    for i in range(1, worksheet.max_column + 1): 
        worksheet.column_dimensions[get_column_letter(i)].width = cell_width 
        
    print('saving...')
    
    workbook.save(excel_path)
    print('success!')

if __name__ == '__main__':
    image_path = "mc.png"
    excel = image_path.split(".")[0] + ".xlsx"
    draw_jpg(image_path, excel)
